import openpyxl
from app_materials.models import Materials
from django.core.management.base import BaseCommand

class Command(BaseCommand):
    help = 'Import data from materials.xlsx into Materials model'

    def handle(self, *args, **kwargs):
        # Path to your materials.xlsx file
        excel_file_path = r'C:\Users\HP\Desktop\mkinfo\app_materials\management\commands\material.xlsx'
        
        # Load the Excel file
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active
        
        # Iterate over the rows and insert data into Materials
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
            material_csr_code = row[0].value
            material_name = row[1].value
            material_measure = row[2].value
            
            # Create a new Materials record
            Materials.objects.create(
                material_csr_code=material_csr_code,
                material_name=material_name,
                material_measure=material_measure
            )
            
        self.stdout.write(self.style.SUCCESS('Material data imported successfully!'))
